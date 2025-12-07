import customtkinter as ctk
from tkinter import messagebox
import threading
from queue import Queue, Empty
from database import add_club, update_club, get_next_club_membership_id, get_club_points_history
from openpyxl import Workbook, load_workbook
from ui_forms import CollapsibleFrame
from utils import bind_mouse_wheel, DateEntry
import os
import shutil
import json
from tkinter import filedialog
from tkinter.ttk import Treeview, Style

class AddClubFrame(ctk.CTkFrame):
    def __init__(self, master, club_data=None, on_save_callback=None, **kwargs):
        super().__init__(master, **kwargs)

        self.club_data = club_data
        self.on_save_callback = on_save_callback

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.entries = {}
        self.multi_attachments = {}

        # Queue for background save operations
        self.save_queue = Queue()
        self.after(100, self._process_save_queue)

        # Queue for background import operations
        self.import_queue = Queue()
        self.after(100, self._process_import_queue)

        # --- Main Scrollable Frame ---
        scrollable_frame = ctk.CTkScrollableFrame(self)
        scrollable_frame.grid(row=0, column=0, sticky="nsew")

        bind_mouse_wheel(scrollable_frame)
        scrollable_frame.grid_columnconfigure(0, weight=1)

        # --- Bulk Import Frame ---
        import_frame = ctk.CTkFrame(scrollable_frame)
        import_frame.pack(fill="x", padx=10, pady=10)
        import_frame.grid_columnconfigure(0, weight=1)

        import_button = ctk.CTkButton(import_frame, text="Import Clubs from Excel", command=self._import_from_excel)
        import_button.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        template_button = ctk.CTkButton(import_frame, text="Download Excel Template", command=self._download_club_excel_template, fg_color="gray")
        template_button.grid(row=0, column=1, padx=5, pady=5, sticky="e")

        # --- Main Club Information ---
        main_info_frame = CollapsibleFrame(scrollable_frame, text="Main Club Information")
        main_info_frame.pack(fill="x", padx=10, pady=5)
        self._create_entry_fields(main_info_frame.content_frame, [
            ("club_membership_id", "Club Membership ID"),
            ("name", "Club Name *"),
            ("address", "Club Address"),
            ("phone", "Phone Number"),
            ("email", "Email"),
            ("classification", "Classification")
        ])
        
        # Make the Club Membership ID field read-only
        self.entries["club_membership_id"].configure(state="disabled")

        # --- Representative Information ---
        rep_info_frame = CollapsibleFrame(scrollable_frame, text="Representative Information")
        rep_info_frame.pack(fill="x", padx=10, pady=5)
        rep_content = rep_info_frame.content_frame
        rep_content.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(rep_content, text="Representative Name:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.entries['representative_name'] = ctk.CTkEntry(rep_content, placeholder_text="Representative Name")
        self.entries['representative_name'].grid(row=0, column=1, padx=10, pady=5, sticky="ew")
        ctk.CTkLabel(rep_content, text="Representative Gender:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.entries['representative_gender'] = ctk.CTkComboBox(rep_content, values=["Male", "Female"])
        self.entries['representative_gender'].set("")
        self.entries['representative_gender'].grid(row=1, column=1, padx=10, pady=5, sticky="ew")

        # --- Points History ---
        points_frame = CollapsibleFrame(scrollable_frame, text="Points History")
        points_frame.pack(fill="x", padx=10, pady=5)
        points_content = points_frame.content_frame
        points_content.grid_columnconfigure(0, weight=1)

        # Input fields for new points entry
        points_input_frame = ctk.CTkFrame(points_content)
        points_input_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=5, pady=5)
        points_input_frame.grid_columnconfigure(1, weight=1)

        self.points_date_entry = DateEntry(points_input_frame, placeholder_text="Date")
        self.points_date_entry.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        self.points_desc_entry = ctk.CTkEntry(points_input_frame, placeholder_text="Description (e.g., Championship Name)")
        self.points_desc_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.points_value_entry = ctk.CTkEntry(points_input_frame, placeholder_text="Points")
        self.points_value_entry.grid(row=0, column=2, padx=5, pady=5, sticky="ew")
        add_points_button = ctk.CTkButton(points_input_frame, text="Add", width=60, command=self._add_points_to_tree)
        add_points_button.grid(row=0, column=3, padx=5, pady=5)

        # Treeview for displaying points history
        self.points_tree = Treeview(points_content, columns=("Date", "Description", "Points"), show="headings", height=5)
        self.points_tree.heading("Date", text="Date")
        self.points_tree.heading("Description", text="Description")
        self.points_tree.heading("Points", text="Points")
        self.points_tree.column("Date", width=100, anchor="center")
        self.points_tree.column("Description", width=300)
        self.points_tree.column("Points", width=80, anchor="center")
        self.points_tree.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        bind_mouse_wheel(self.points_tree)

        # Action buttons for tree
        points_tree_actions_frame = ctk.CTkFrame(points_content, fg_color="transparent")
        points_tree_actions_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=5)
        remove_points_button = ctk.CTkButton(points_tree_actions_frame, text="Remove Selected", command=self._remove_points_from_tree)
        remove_points_button.pack(side="left")

        # --- Subscription & Affiliation ---
        sub_info_frame = CollapsibleFrame(scrollable_frame, text="Subscription & Affiliation")
        sub_info_frame.pack(fill="x", padx=10, pady=5)
        self._create_entry_fields(sub_info_frame.content_frame, [
            ("affiliation_date", "Affiliation Date *"),
            ("subscription_expiry_date", "Subscription Expiry Date *"),
            ("club_subscription_fee", "Club Subscription Fee"),
            ("admin_subscription_fee", "Admin Subscription Fee")
        ])

        # --- Total Points Display ---
        total_points_frame = ctk.CTkFrame(scrollable_frame, fg_color=("gray85", "gray25"))
        total_points_frame.pack(fill="x", padx=10, pady=10)
        total_points_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(total_points_frame, text="Total Points:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10, pady=5)
        self.total_points_label = ctk.CTkLabel(total_points_frame, text="0", font=ctk.CTkFont(weight="bold", size=16))
        self.total_points_label.pack(side="left", padx=10, pady=5)

        # --- Attachments ---
        attachment_frame = CollapsibleFrame(scrollable_frame, text="Attachments")
        attachment_frame.pack(fill="x", padx=10, pady=5)
        self.attachment_content = attachment_frame.content_frame
        self.attachment_content.grid_columnconfigure(0, weight=1)

        self.multi_attachment_fields_meta = {
            'federation_license': ("Federation License", 0),
            'olympic_license': ("Olympic Committee License", 1),
            'payment_receipt': ("Payment Receipts", 2)
        }
        for key, (text, row) in self.multi_attachment_fields_meta.items():
            frame = self._create_multi_attachment_field(self.attachment_content, text, key)
            frame.grid(row=row, column=0, sticky="nsew", padx=5, pady=5)
            if key not in self.multi_attachments:
                self.multi_attachments[key] = {}
            self.multi_attachments[key]['frame'] = frame

        # --- Action Buttons ---
        button_frame = ctk.CTkFrame(self)
        button_frame.grid(row=1, column=0, padx=10, pady=20, sticky="ew")
        button_frame.grid_columnconfigure(1, weight=1)

        # Don't show clear button in edit mode
        if not self.club_data:
            clear_button = ctk.CTkButton(button_frame, text="Clear Form", command=self._clear_form, fg_color="gray")
            clear_button.grid(row=0, column=0, padx=(0, 10))

        button_text = "Update Club" if self.club_data else "Save Club"
        self.save_button = ctk.CTkButton(button_frame, text=button_text, command=self._save_club, height=40)
        self.save_button.grid(row=0, column=1, sticky="ew")

        # If in edit mode, populate the form with existing data
        if self.club_data:
            self.populate_form()
        else:
            self.set_next_club_id()

    def _process_save_queue(self):
        """Processes results from the background save worker thread."""
        try:
            result_type, data = self.save_queue.get_nowait()
            self.save_button.configure(state="normal", text="Update Club" if self.club_data else "Save Club")

            if result_type == "save_success":
                club_name = data
                messagebox.showinfo("Success", f"Club '{club_name}' has been saved/updated successfully.")
                
                if self.on_save_callback:
                    self.on_save_callback()

                if isinstance(self.master, ctk.CTkToplevel):
                    self.master.destroy()
                else:
                    self._clear_form()

                # Refresh sibling frames that depend on the club list
                if hasattr(self.master.master, 'add_member_frame'):
                    self.master.master.add_member_frame.update_club_list()
                if hasattr(self.master.master, 'reports_frame'):
                    self.master.master.reports_frame.update_club_filter()

            elif result_type == "save_error":
                error_message = data
                messagebox.showerror("Database Error", f"An error occurred while saving the club: {error_message}")
        except Empty:
            pass
        finally:
            self.after(100, self._process_save_queue)

    def _process_import_queue(self):
        """Processes results from the background Excel import worker thread."""
        try:
            result_type, data = self.import_queue.get_nowait()
            if result_type == "import_finished":
                success_count, fail_count, failed_items = data
                summary_message = f"Club import complete.\n\nSuccessfully imported: {success_count} clubs.\nFailed to import: {fail_count} clubs."
                if failed_items:
                    summary_message += f"\n\nFailed Club Names/IDs:\n" + ",\n".join(failed_items)
                
                messagebox.showinfo("Import Summary", summary_message)
                
                # Refresh club lists in other frames
                if hasattr(self.master.master, 'add_member_frame'):
                    self.master.master.add_member_frame.update_club_list()
                if hasattr(self.master.master, 'reports_frame'):
                    self.master.master.reports_frame.update_club_filter()

            elif result_type == "import_error":
                error_message = data
                messagebox.showerror("Import Error", f"An error occurred during the import process: {error_message}")
        except Empty:
            pass
        finally:
            self.after(100, self._process_import_queue)

    def _import_from_excel(self):
        filepath = filedialog.askopenfilename(
            title="Select Excel File for Clubs",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not filepath:
            return

        messagebox.showinfo("Importing", "Importing clubs from Excel in the background. You will be notified upon completion.")
        
        thread = threading.Thread(target=self._import_clubs_from_excel_worker, args=(filepath,), daemon=True)
        thread.start()

    def _import_clubs_from_excel_worker(self, filepath):
        """Worker function to handle reading the Excel file and inserting clubs into DB."""
        try:
            workbook = load_workbook(filepath, data_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            success_count, fail_count, failed_items = 0, 0, []

            for row_index in range(2, sheet.max_row + 1):
                row_dict = {}
                try:
                    row_values = [cell.value for cell in sheet[row_index]]
                    if not any(row_values): continue
                    row_dict = dict(zip(headers, row_values))
                    club_data = {}
                    if not row_dict.get('club_membership_id'):
                        club_data['club_membership_id'] = get_next_club_membership_id()
                    if not row_dict.get('name'): raise ValueError("Club 'name' is missing.")
                    for key in headers:
                        if key in row_dict and row_dict[key] is not None:
                            club_data[key] = row_dict[key]
                    club_data.setdefault('points', 0)
                    club_data.setdefault('attachments_data', {})
                    add_club(club_data)
                    success_count += 1
                except Exception as e:
                    fail_count += 1
                    failed_items.append(str(row_dict.get('name', f'Row {row_index}')))
                    print(f"Failed to import club at row {row_index}: {e}")
            
            self.import_queue.put(("import_finished", (success_count, fail_count, failed_items)))
        except Exception as e:
            self.import_queue.put(("import_error", str(e)))

    def _create_entry_fields(self, master, fields):
        master.grid_columnconfigure(1, weight=1)
        for i, (key, placeholder) in enumerate(fields):
            ctk.CTkLabel(master, text=f"{placeholder.replace('*', '').strip()}:").grid(row=i, column=0, padx=10, pady=5, sticky="w")
            if "Date" in placeholder or "Expiry" in placeholder:
                entry = DateEntry(master, placeholder_text=placeholder)
            else:
                entry = ctk.CTkEntry(master, placeholder_text=placeholder)

            entry.grid(row=i, column=1, padx=10, pady=5, sticky="ew")
            self.entries[key] = entry

    def _create_multi_attachment_field(self, master, text, category_key):
        main_frame = ctk.CTkFrame(master)
        main_frame.grid_columnconfigure(0, weight=1)
        header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        header_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=(5, 0))
        header_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(header_frame, text=text).grid(row=0, column=0, sticky="w")
        upload_button = ctk.CTkButton(header_frame, text="Upload Files", width=120, command=lambda: self._upload_multiple_files(category_key))
        upload_button.grid(row=0, column=1, sticky="e")
        tree_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)
        tree = Treeview(tree_frame, columns=("filename",), show="headings", height=3)
        tree.heading("filename", text="File Name")
        tree.column("filename", anchor="w")
        tree.grid(row=0, column=0, sticky="nsew")
        bind_mouse_wheel(tree)
        scrollbar = ctk.CTkScrollbar(tree_frame, command=tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=scrollbar.set)
        if category_key not in self.multi_attachments:
            self.multi_attachments[category_key] = {}
        self.multi_attachments[category_key]['tree'] = tree
        remove_button = ctk.CTkButton(main_frame, text="Remove Selected File", command=lambda: self._remove_attachment_from_tree(category_key))
        remove_button.grid(row=2, column=0, sticky="w", padx=5, pady=(0, 5))
        return main_frame

    def _upload_multiple_files(self, category_key):
        filepaths = filedialog.askopenfilenames(title=f"Select File(s) for {category_key.replace('_', ' ').title()}")
        if not filepaths: return
        club_id = self.entries['club_membership_id'].get()
        if not club_id:
            messagebox.showwarning("Warning", "Please enter a Club Membership ID before uploading files.")
            return
        safe_club_id = "".join(c for c in club_id if c.isalnum() or c in ('-', '_')).rstrip()
        club_assets_dir = os.path.join("assets", "club_files", safe_club_id)
        os.makedirs(club_assets_dir, exist_ok=True)
        tree = self.multi_attachments[category_key]['tree']
        for filepath in filepaths:
            destination_path = os.path.join(club_assets_dir, f"{category_key}_{os.path.basename(filepath)}")
            
            counter = 1
            while os.path.exists(destination_path):
                name, ext = os.path.splitext(f"{category_key}_{os.path.basename(filepath)}")
                destination_path = os.path.join(club_assets_dir, f"{name}_{counter}{ext}")
                counter += 1

            shutil.copy(filepath, destination_path)
            tree.insert("", "end", iid=destination_path, values=(os.path.basename(destination_path),))

    def _remove_attachment_from_tree(self, category_key):
        tree = self.multi_attachments[category_key]['tree']
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select a file to remove.")
            return
        if messagebox.askyesno("Confirm Deletion", f"Are you sure you want to remove {len(selected_items)} selected file(s) from the list and delete them from disk? This action cannot be undone."):
            for item_id in selected_items:
                filepath = item_id
                try:
                    if os.path.exists(filepath): os.remove(filepath)
                    tree.delete(item_id)
                except Exception as e:
                    messagebox.showerror("Error", f"Could not delete file: {e}")

    def set_next_club_id(self):
        """Fetches the next club ID and displays it in the form."""
        next_id = get_next_club_membership_id()
        self.entries['club_membership_id'].configure(state="normal")
        self.entries['club_membership_id'].delete(0, "end")
        self.entries['club_membership_id'].insert(0, next_id)
        self.entries['club_membership_id'].configure(state="disabled")

    def _add_points_to_tree(self):
        date = self.points_date_entry.get()
        description = self.points_desc_entry.get()
        points_str = self.points_value_entry.get()
        if not all([date, description, points_str]):
            messagebox.showwarning("Input Missing", "Please fill in Date, Description, and Points.")
            return
        try:
            points = int(points_str)
            self.points_tree.insert("", "end", values=(date, description, points))
            self.points_date_entry.delete(0, "end")
            self.points_desc_entry.delete(0, "end")
            self.points_value_entry.delete(0, "end")
            self._update_total_points()
        except ValueError:
            messagebox.showerror("Invalid Input", "Points must be a valid number.")

    def _remove_points_from_tree(self):
        selected_item = self.points_tree.selection()
        if selected_item:
            self.points_tree.delete(selected_item)
            self._update_total_points()

    def _update_total_points(self):
        total = sum(int(self.points_tree.item(item, "values")[2]) for item in self.points_tree.get_children())
        self.total_points_label.configure(text=str(total))

    def populate_form(self):
        """Fills the form with existing club data for editing."""
        if not self.club_data:
            return
        
        for key, widget in self.entries.items():
            value = self.club_data.get(key)
            if value is None:
                value = ''
            if isinstance(widget, ctk.CTkEntry):
                widget.insert(0, str(value))
            elif isinstance(widget, DateEntry):
                widget.set(str(value))
            elif isinstance(widget, ctk.CTkComboBox):
                widget.set(str(value))
        
        # Populate points history
        points_history = get_club_points_history(self.club_data['id'])
        for item in self.points_tree.get_children():
            self.points_tree.delete(item)
        for record in points_history:
            self.points_tree.insert("", "end", values=(record['date'], record['description'], record['points']))
        self._update_total_points()

        # Populate attachments
        try:
            attachments_data = json.loads(self.club_data.get('attachments_data', '{}'))
            for key, attachment_info in self.multi_attachments.items():
                if key in attachments_data:
                    tree = attachment_info.get('tree')
                    if tree:
                        file_paths = attachments_data[key]
                        if isinstance(file_paths, list):
                            for path in file_paths:
                                if os.path.exists(path):
                                    tree.insert("", "end", iid=path, values=(os.path.basename(path),))
        except (json.JSONDecodeError, TypeError):
            print("Could not parse or populate club attachments.")

    def _save_club_worker(self, club_data, is_update, club_id=None):
        """Worker function to save or update club data in the background."""
        try:
            if is_update:
                update_club(club_id, club_data)
            else:
                add_club(club_data)
            self.save_queue.put(("save_success", club_data['name']))
        except Exception as e:
            self.save_queue.put(("save_error", str(e)))

    def _save_club(self):
        club_data = {key: entry.get() for key, entry in self.entries.items()}

        # --- Validation ---
        required_fields_map = {
            "name": "Club Name",
            "affiliation_date": "Affiliation Date",
            "subscription_expiry_date": "Subscription Expiry Date"
        }
        for key, label in required_fields_map.items():
            if not club_data.get(key):
                messagebox.showerror("Validation Error", f"'{label}' is a required field.")
                return

        # Convert fee fields to float, default to None if empty
        try:
            fee_fields = ['club_subscription_fee', 'admin_subscription_fee']
            for field in fee_fields:
                value = self.entries[field].get()
                club_data[field] = float(value) if value else None
        except ValueError as e:
            messagebox.showerror("Input Error", f"Please enter valid numbers for fees. Error: {e}")
            return

        # Collect points history
        points_history = []
        for item_id in self.points_tree.get_children():
            values = self.points_tree.item(item_id, 'values')
            points_history.append({'date': values[0], 'description': values[1], 'points': values[2]})
        club_data['points_history'] = points_history

        # Collect attachment data
        club_data['attachments_data'] = {}
        for key, attachment_info in self.multi_attachments.items():
            tree = attachment_info.get('tree')
            if tree:
                file_paths = tree.get_children() # Item IDs are the full paths
                club_data['attachments_data'][key] = file_paths
        
        # Disable button and show message
        self.save_button.configure(state="disabled", text="Saving...")
        self.update_idletasks()

        is_update = self.club_data is not None
        club_id = self.club_data['id'] if is_update else None

        thread = threading.Thread(target=self._save_club_worker, args=(club_data, is_update, club_id), daemon=True)
        thread.start()

    def _download_club_excel_template(self):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")],
            initialfile="pkf_clubs_import_template.xlsx"
        )
        if not filepath:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Clubs Import"
        headers = [
            "club_membership_id", "name", "representative_name", "representative_gender",
            "address", "phone", "email", "classification", "points",
            "affiliation_date", "subscription_expiry_date",
            "club_subscription_fee", "admin_subscription_fee"
        ]
        ws.append(headers)
        wb.save(filepath)
        messagebox.showinfo("Success", f"Club import template saved to:\n{filepath}")

    def _clear_form(self):
        for entry in self.entries.values():
            if isinstance(entry, ctk.CTkComboBox):
                entry.set("")
            else:
                entry.delete(0, 'end')
            if isinstance(entry, DateEntry):
                entry.delete(0, 'end')
        
        # Clear points history tree
        for item in self.points_tree.get_children():
            self.points_tree.delete(item)
        self._update_total_points()

        # Clear multi-attachment trees
        for attachment_info in self.multi_attachments.values():
            tree = attachment_info.get('tree')
            if tree:
                for item in tree.get_children():
                    tree.delete(item)
        
        self.set_next_club_id()