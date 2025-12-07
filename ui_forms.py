import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading
from queue import Queue, Empty
from tkinter.ttk import Treeview, Style
import os
import shutil
from database import add_member, update_member, get_belts, get_achievements, get_all_clubs, get_next_pkf_id, find_member_by_name
from openpyxl import Workbook, load_workbook
from utils import bind_mouse_wheel, DateEntry, get_eligible_categories, calculate_age
import json


class CollapsibleFrame(ctk.CTkFrame):
    """A collapsible frame widget for customtkinter."""
    def __init__(self, master, text="", **kwargs):
        super().__init__(master, **kwargs)
        self.grid_columnconfigure(0, weight=1)
        self.text = text
        self.is_collapsed = False

        self.header_button = ctk.CTkButton(self, text=f"−  {self.text}", command=self.toggle, anchor="w", fg_color="transparent", text_color=("black", "white"), hover=False)
        self.header_button.grid(row=0, column=0, sticky="ew")

        self.content_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.content_frame.grid(row=1, column=0, padx=5, sticky="nsew")

    def toggle(self):
        self.is_collapsed = not self.is_collapsed
        if self.is_collapsed:
            self.content_frame.grid_remove()
            self.header_button.configure(text=f"+  {self.text}")
        else:
            self.content_frame.grid()
            self.header_button.configure(text=f"−  {self.text}")

class AddMemberFrame(ctk.CTkFrame):
    def __init__(self, master, member_data=None, on_save_callback=None, **kwargs):
        super().__init__(master, **kwargs)

        self.member_data = member_data # If not None, we are in "edit mode"
        self.on_save_callback = on_save_callback # To refresh search results
        self.specific_widgets = {}
        
        self.photo_path_var = ctk.StringVar()
        self.multi_attachments = {} # Will hold widgets and data for multi-upload fields

        # Queue for background tasks like Excel import
        self.import_queue = Queue()
        self.after(100, self._process_import_queue)

        # Queue for background save operations
        self.save_queue = Queue()
        self.after(100, self._process_save_queue)

        # Queue for loading club list
        self.club_queue = Queue()
        self.after(100, self._process_club_queue)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # --- Main Scrollable Frame ---
        scrollable_frame = ctk.CTkScrollableFrame(self)
        scrollable_frame.grid(row=0, column=0, sticky="nsew")

        bind_mouse_wheel(scrollable_frame)

        scrollable_frame.grid_columnconfigure(0, weight=1)

        # --- Bulk Import Frame ---
        import_frame = ctk.CTkFrame(scrollable_frame)
        import_frame.pack(fill="x", padx=10, pady=10)
        import_frame.grid_columnconfigure(0, weight=1)

        import_button = ctk.CTkButton(import_frame, text="Import Members from Excel", command=self._import_from_excel)
        import_button.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        template_button = ctk.CTkButton(import_frame, text="Download Excel Template", command=self._download_excel_template, fg_color="gray")
        template_button.grid(row=0, column=1, padx=5, pady=5, sticky="e")


        # --- Data Variables ---
        self.entries = {}

        # --- Basic Information ---
        basic_frame = CollapsibleFrame(scrollable_frame, text="Basic Information")
        basic_frame.pack(fill="x", padx=10, pady=5)
        basic_content = basic_frame.content_frame

        # Fields before DOB
        fields_before_dob = [
            ("pkf_id", "Membership No."), ("id_number", "ID No."),
            ("full_name", "Full Name (English)"), ("full_name_ar", "Full Name (Arabic)"),
            ("passport_number", "Passport No."),
            ("passport_expiry_date", "Passport Expiry"),
            ("place_of_birth", "Place of Birth"),
        ]
        self._create_entry_fields(basic_content, fields_before_dob, start_row=0)

        # --- DOB and Age ---
        dob_row = len(fields_before_dob)
        ctk.CTkLabel(basic_content, text="Date of Birth:").grid(row=dob_row, column=0, padx=10, pady=5, sticky="w")

        dob_age_frame = ctk.CTkFrame(basic_content, fg_color="transparent")
        dob_age_frame.grid(row=dob_row, column=1, padx=10, pady=5, sticky="ew")
        dob_age_frame.grid_columnconfigure(0, weight=1) # DOB entry
        dob_age_frame.grid_columnconfigure(2, weight=0) # Age entry

        dob_entry = DateEntry(dob_age_frame, placeholder_text="YYYY-MM-DD", command=self._update_age_and_categories)
        dob_entry.grid(row=0, column=0, sticky="ew")
        self.entries['dob'] = dob_entry

        ctk.CTkLabel(dob_age_frame, text="Age:").grid(row=0, column=1, padx=(10, 5), sticky="w")

        age_entry = ctk.CTkEntry(dob_age_frame, width=50, state="disabled")
        age_entry.grid(row=0, column=2, sticky="w")
        self.entries['age'] = age_entry

        # --- Fields after DOB ---
        fields_after_dob = [
            ("profession", "Profession"), ("education", "Education"),
            ("address", "Address"), ("phone", "Phone"), ("email", "Email"),
            ("federation_join_date", "Federation Join Date (YYYY-MM-DD)")
        ]
        self._create_entry_fields(basic_content, fields_after_dob, start_row=dob_row + 1)

        # Bind event to Arabic name entry for auto-finding member ID
        self.entries['full_name_ar'].bind("<FocusOut>", self._check_existing_member)
        self.entries['full_name_ar'].bind("<Return>", self._check_existing_member)
        
        # --- Add Gender Combobox ---
        gender_row = dob_row + 1 + len(fields_after_dob)
        ctk.CTkLabel(basic_content, text="Gender:").grid(row=gender_row, column=0, padx=10, pady=5, sticky="w")
        gender_combo = ctk.CTkComboBox(basic_content, values=["Male", "Female"])
        gender_combo.grid(row=gender_row, column=1, padx=10, pady=5, sticky="ew")
        self.entries['gender'] = gender_combo
        # Bind events for dob, gender, and weight to update competition categories
        self.entries['gender'].bind("<<ComboboxSelected>>", self._update_age_and_categories)
        self.entries['dob'].entry.bind("<FocusOut>", self._update_age_and_categories)
        self.entries['dob'].entry.bind("<Return>", self._update_age_and_categories)


        # --- Club Information ---
        club_frame = CollapsibleFrame(scrollable_frame, text="Club Selection")
        club_frame.pack(fill="x", padx=10, pady=5)
        club_content = club_frame.content_frame
        club_content.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(club_content, text="Club:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.club_combobox = ctk.CTkComboBox(club_content, values=["Loading clubs..."], state="disabled")
        self.club_combobox.grid(row=0, column=1, padx=10, pady=5, sticky="ew")
        self.clubs_map = {} # To map club names to their IDs
        self.update_club_list()

        # --- Financial & Role ---
        finance_frame = CollapsibleFrame(scrollable_frame, text="Financial & Role")
        finance_frame.pack(fill="x", padx=10, pady=5)
        finance_content = finance_frame.content_frame
        finance_content.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(finance_content, text="Role:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.role_menu = ctk.CTkOptionMenu(finance_content, values=["Player", "Coach", "Referee", "Admin"], command=self._update_role_fields)
        self.role_menu.grid(row=0, column=1, columnspan=3, padx=10, pady=5, sticky="ew")
        self.entries['role'] = self.role_menu

        self._create_entry_fields(finance_content, [
            ("subscription_fee", "Subscription Fee"), # New
            ("expiry_date", "Subscription Expiry") # New
        ], start_row=1)

        # --- Belt Information ---
        belt_frame = CollapsibleFrame(scrollable_frame, text="Belt Information")
        belt_frame.pack(fill="x", padx=10, pady=5)
        self._create_entry_fields(belt_frame.content_frame, [
            ("current_belt", "Current Belt"), # New
            ("current_belt_date", "Date Obtained") # New
        ])
        
        # --- Belt History Table ---
        belt_history_frame = ctk.CTkFrame(belt_frame.content_frame)
        belt_history_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=5, pady=10)
        belt_history_frame.grid_columnconfigure(0, weight=1)


        # Input fields for new belt
        self.belt_name_entry = ctk.CTkEntry(belt_history_frame, placeholder_text="Belt")
        self.belt_name_entry.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
        self.belt_date_entry = DateEntry(belt_history_frame, placeholder_text="Date")
        self.belt_date_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        self.belt_source_entry = ctk.CTkEntry(belt_history_frame, placeholder_text="Source") # New
        self.belt_source_entry.grid(row=1, column=2, padx=5, pady=5, sticky="ew")
        
        add_belt_button = ctk.CTkButton(belt_history_frame, text="Add", width=60, command=self._add_belt_to_tree)
        add_belt_button.grid(row=1, column=3, padx=5, pady=5)

        # Treeview for displaying belts
        style = Style()
        style.theme_use("default")
        style.configure("Treeview", background="#2B2B2B", foreground="white", fieldbackground="#2B2B2B", borderwidth=0)
        style.map('Treeview', background=[('selected', '#8A2BE2')])
        style.configure("Treeview.Heading", background="#565B5E", foreground="white", relief="flat")
        ctk.CTkLabel(belt_history_frame, text="Belt History", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=4, pady=(5,10))
        style.map("Treeview.Heading", background=[('active', '#343638')])

        self.belt_tree = Treeview(belt_history_frame, columns=("Belt", "Date", "Source"), show="headings", height=4)
        self.belt_tree.heading("Belt", text="Belt")
        self.belt_tree.heading("Date", text="Date")
        self.belt_tree.heading("Source", text="Source")
        self.belt_tree.column("Belt", width=120)
        self.belt_tree.column("Date", width=120)
        self.belt_tree.column("Source", width=150)
        self.belt_tree.grid(row=2, column=0, columnspan=4, sticky="nsew", padx=5, pady=5)
        bind_mouse_wheel(self.belt_tree)


        remove_belt_button = ctk.CTkButton(belt_history_frame, text="Remove Selected", command=self._remove_belt_from_tree)
        remove_belt_button.grid(row=3, column=0, columnspan=4, padx=5, pady=5, sticky="w")


        # --- Category History ---
        category_frame = CollapsibleFrame(scrollable_frame, text="Category History")
        category_frame.pack(fill="x", padx=10, pady=5)
        category_content = category_frame.content_frame
        category_content.grid_columnconfigure(1, weight=1)
        category_content.grid_columnconfigure(2, weight=1)

        ctk.CTkLabel(category_content, text="Category", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5)
        ctk.CTkLabel(category_content, text="Start Date", font=ctk.CTkFont(weight="bold")).grid(row=0, column=1, padx=10, pady=5)
        ctk.CTkLabel(category_content, text="End Date", font=ctk.CTkFont(weight="bold")).grid(row=0, column=2, padx=10, pady=5)

        categories = ["Player", "Coach", "Referee", "Admin"]
        for i, category in enumerate(categories):
            key_prefix = category.lower()
            ctk.CTkLabel(category_content, text=f"{category}:").grid(row=i + 1, column=0, padx=10, pady=5, sticky="w") # New
            start_entry = DateEntry(category_content, placeholder_text="Start Date")
            start_entry.grid(row=i + 1, column=1, padx=10, pady=5, sticky="ew")
            self.entries[f'{key_prefix}_start_date'] = start_entry
            end_entry = DateEntry(category_content, placeholder_text="End Date")
            end_entry.grid(row=i + 1, column=2, padx=10, pady=5, sticky="ew")
            self.entries[f'{key_prefix}_end_date'] = end_entry
        
        # --- Role-Specific Fields ---
        self.specific_fields_main_frame = CollapsibleFrame(scrollable_frame, text="Specialization Details")
        self.specific_fields_main_frame.pack(fill="x", padx=10, pady=5)
        specific_content = self.specific_fields_main_frame.content_frame
        specific_content.grid_columnconfigure(1, weight=1)

        self._create_player_fields(specific_content)
        self._create_coach_fields(specific_content)
        self._create_referee_fields(specific_content)
        self._create_admin_fields(specific_content)

        # --- Notes ---
        notes_frame = CollapsibleFrame(scrollable_frame, text="Notes")
        notes_frame.pack(fill="x", padx=10, pady=5)
        notes_content = notes_frame.content_frame
        notes_content.grid_columnconfigure(0, weight=1)
        self.notes_textbox = ctk.CTkTextbox(notes_content, height=100)
        self.notes_textbox.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        self.entries['notes'] = self.notes_textbox

        # --- Attachments ---
        attachment_frame = CollapsibleFrame(scrollable_frame, text="Attachments")
        attachment_frame.pack(fill="x", padx=10, pady=5)
        self.attachment_content = attachment_frame.content_frame
        self.attachment_content.grid_columnconfigure(0, weight=1)

        # --- Personal Photo (Single Upload) ---
        photo_main_frame = ctk.CTkFrame(self.attachment_content)
        photo_main_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        photo_main_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(photo_main_frame, text="Personal Photo:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        photo_upload_frame = ctk.CTkFrame(photo_main_frame, fg_color="transparent")
        photo_upload_frame.grid(row=0, column=1, sticky="ew")
        photo_upload_frame.grid_columnconfigure(1, weight=1)
        self.photo_upload_btn = ctk.CTkButton(photo_upload_frame, text="Upload", width=100, command=self._upload_personal_photo)
        self.photo_upload_btn.grid(row=0, column=0, sticky="w")
        self.photo_path_label = ctk.CTkLabel(photo_upload_frame, text="No file selected.", text_color="gray")
        self.photo_path_label.grid(row=0, column=1, padx=10, sticky="w")

        # --- Multi-upload fields ---
        self.multi_attachment_fields_meta = {
            'identity_docs': ("Identity Documents (Passport, ID, etc.)", 1),
            'belt_certs': ("Belt Certificates", 2),
            'coach_certs': ("Coaching Certificates", 3),
            'referee_certs': ("Refereeing Certificates", 4),
            'payment_receipts': ("Payment Receipts", 5)
        }
        for key, (text, row) in self.multi_attachment_fields_meta.items():
            frame = self._create_multi_attachment_field(self.attachment_content, text, key)
            frame.grid(row=row, column=0, sticky="nsew", padx=5, pady=5)
            self.multi_attachments[key]['frame'] = frame
        # --- Button Text ---
        button_text = "Update Member" if self.member_data else "Save Member"

        # --- Save Button ---
        self.save_button = ctk.CTkButton(self, text=button_text, command=self._save_member, height=40)
        self.save_button.grid(row=1, column=0, padx=10, pady=20, sticky="ew")

        # Initial field setup

        if self.member_data:
            self.populate_form()
        else:
            # This is a new member, so set the next PKF ID and make it read-only
            self.set_next_pkf_id()
        
        # This needs to be called after populating form in edit mode
        self._update_role_fields(self.role_menu.get())

    def _process_club_queue(self):
        try:
            result_type, data = self.club_queue.get_nowait()
            if result_type == "club_list":
                all_clubs = data
                self.clubs_map = {club['name']: club['id'] for club in all_clubs}
                club_names = sorted(self.clubs_map.keys())
                self.club_combobox.configure(values=club_names if club_names else ["No clubs available"], state="readonly")
                self.club_combobox.set("Select a club" if club_names else "No clubs available")
                # If in edit mode, re-populate the club selection after the list is loaded
                if self.member_data:
                    self._populate_club_selection()
            elif result_type == "club_list_error":
                self.club_combobox.configure(values=["Error loading clubs"], state="disabled")
        except Empty:
            pass
        finally:
            self.after(100, self._process_club_queue)

    def _update_club_list_worker(self):
        try:
            all_clubs = get_all_clubs()
            self.club_queue.put(("club_list", all_clubs))
        except Exception as e:
            self.club_queue.put(("club_list_error", str(e)))

    def update_club_list(self):
        """Fetches the latest list of clubs from the database asynchronously."""
        self.club_combobox.configure(values=["Loading clubs..."], state="disabled")
        threading.Thread(target=self._update_club_list_worker, daemon=True).start()

    def set_next_pkf_id(self):
        """Fetches the next PKF ID and sets it in the form, making the field read-only."""
        self.entries['pkf_id'].configure(state="normal") # Enable to modify
        next_id = get_next_pkf_id()
        self.entries['pkf_id'].delete(0, 'end')
        self.entries['pkf_id'].insert(0, next_id)
        self.entries['pkf_id'].configure(state="disabled") # Disable again

    def _prefill_for_new_role(self, existing_member_data):
        """
        Populates the form with existing member's general data
        to allow adding a new role for the same PKF ID.
        """
        messagebox.showinfo(
            "عضو موجود",
            f"تم العثور على عضو بالاسم العربي '{existing_member_data.get('full_name_ar')}' "
            f"ورقم العضوية '{existing_member_data.get('pkf_id')}'.\n\n"
            "تم تعبئة النموذج بمعلوماته الأساسية. يمكنك الآن تحديد دور جديد وتفاصيل أخرى لإضافة إدخال جديد لهذا العضو."
        )

        # Clear the form first to ensure all role-specific fields are reset
        # We pass True to indicate it's a prefill, so it doesn't reset PKF ID
        self.clear_form(is_prefill=True)

        # Populate common fields
        common_fields = [
            'pkf_id', 'full_name', 'full_name_ar', 'id_number', 'dob', 'gender',
            'phone', 'email', 'passport_number', 'passport_expiry_date'
        ]
        for key in common_fields:
            value = existing_member_data.get(key)
            if value is not None:
                widget = self.entries.get(key)
                if widget:
                    if isinstance(widget, ctk.CTkEntry):
                        widget.insert(0, str(value))
                    elif isinstance(widget, DateEntry):
                        widget.set(str(value))
                    elif isinstance(widget, ctk.CTkComboBox):
                        widget.set(str(value))

        # Special handling for photo_path_var
        photo_path = existing_member_data.get('photo_path')
        if photo_path:
            self.photo_path_var.set(photo_path)
            self.photo_path_label.configure(text=os.path.basename(photo_path), text_color=("black", "white"))
        else:
            self.photo_path_var.set("")
            self.photo_path_label.configure(text="No file selected.", text_color="gray")

        # Set the existing PKF ID and disable it
        self.entries['pkf_id'].configure(state="normal")
        self.entries['pkf_id'].delete(0, 'end')
        self.entries['pkf_id'].insert(0, existing_member_data['pkf_id'])
        self.entries['pkf_id'].configure(state="disabled")

        self._populate_club_selection(existing_member_data)

    def _check_existing_member(self, event=None):
        """
        Checks if a member with the entered Arabic name already exists.
        This is intended for "Add New Member" mode to prevent duplicates.
        """
        # This check is only relevant when adding a new member, not editing.
        if self.member_data:
            return

        name_ar = self.entries['full_name_ar'].get()
        if not name_ar.strip():
            return

        existing_member = find_member_by_name(name_ar)

        if existing_member:
            self._prefill_for_new_role(existing_member)

    def _add_belt_to_tree(self):
        belt = self.belt_name_entry.get()
        date = self.belt_date_entry.get()
        source = self.belt_source_entry.get()
        if belt and date and source:
            self.belt_tree.insert("", "end", values=(belt, date, source))
            self.belt_name_entry.delete(0, "end")
            self.belt_date_entry.delete(0, "end")
            self.belt_source_entry.delete(0, "end")
        else:
            messagebox.showwarning("Warning", "Please fill all three belt fields (Belt, Date, Source).")

    def _remove_belt_from_tree(self):
        selected_item = self.belt_tree.selection()
        if selected_item:
            self.belt_tree.delete(selected_item)
        else:
            messagebox.showwarning("Warning", "Please select a belt from the history table to remove.")

    def _update_age_and_categories(self, event=None):
        """
        Calculates and updates age, and eligible Kumite and Kata categories
        based on DOB, Gender, and Weight.
        """
        print(f"DEBUG: _update_age_and_categories called by event: {event}")
        dob_str = self.entries['dob'].get()

        # --- Age Calculation ---
        age = calculate_age(dob_str)
        age_text = str(age) if age is not None else ""
        if 'age' in self.entries:
            age_entry = self.entries['age']
            age_entry.configure(state="normal")
            age_entry.delete(0, 'end')
            age_entry.insert(0, age_text)
            age_entry.configure(state="disabled")

        gender = self.entries['gender'].get()
        
        # Ensure 'Player' specific widgets are initialized before accessing
        if 'Player' not in self.specific_widgets or 'weight' not in self.specific_widgets['Player']['widgets']:
            print("DEBUG: Player specific widgets not initialized or visible. Returning.")
            return # Player fields not yet created or visible

        weight_str = self.specific_widgets['Player']['widgets']['weight'][1].get()
        print(f"DEBUG: dob_str='{dob_str}', gender='{gender}', weight_str='{weight_str}'")

        # --- New logic for populating checkboxes ---
        player_widgets = self.specific_widgets['Player']['widgets']

        # Clear previous checkboxes
        for widget in self.eligible_categories_frame.winfo_children():
            widget.destroy()
        if 'category_checkboxes' in player_widgets:
            player_widgets['category_checkboxes'].clear()

        if not dob_str or not gender or not weight_str:
            # If essential info is missing, show a message and stop.
            print("DEBUG: Missing essential info. Clearing categories.")
            no_cat_label = ctk.CTkLabel(self.eligible_categories_frame, text="Enter DOB, Gender, and Weight to see categories.")
            no_cat_label.pack(pady=10)
            self._update_selected_categories_display() # This will clear the display
            return

        try:
            weight_kg = float(weight_str)
            print(f"DEBUG: weight_kg={weight_kg}")
        except ValueError:
            # messagebox.showwarning("Invalid Input", "Please enter a valid number for Weight (kg).") # Too many popups
            print(f"DEBUG: Invalid weight_str='{weight_str}'. Returning.")
            no_cat_label = ctk.CTkLabel(self.eligible_categories_frame, text="Invalid weight. Please enter a number.")
            no_cat_label.pack(pady=10)
            self._update_selected_categories_display() # This will clear the display
            return

        eligible_kumite, eligible_kata = get_eligible_categories(dob_str, gender, weight_kg)
        print(f"DEBUG: Eligible Kumite: {eligible_kumite}, Eligible Kata: {eligible_kata}")

        all_eligible_categories = sorted(list(set(eligible_kata + eligible_kumite)))

        if not all_eligible_categories:
            no_cat_label = ctk.CTkLabel(self.eligible_categories_frame, text="No eligible categories found.")
            no_cat_label.pack(pady=10)
        else:
            for category_name in all_eligible_categories:
                checkbox = ctk.CTkCheckBox(
                    self.eligible_categories_frame,
                    text=category_name,
                    command=self._update_selected_categories_display
                )
                checkbox.pack(anchor="w", padx=10, pady=2)
                player_widgets['category_checkboxes'][category_name] = checkbox

        # Also clear/update the selected display
        self._update_selected_categories_display()

    def _process_import_queue(self):
        """Processes results from the background Excel import worker thread."""
        try:
            result_type, data = self.import_queue.get_nowait()
            if result_type == "import_finished":
                success_count, fail_count, failed_ids = data
                summary_message = f"Import complete.\n\nSuccessfully imported: {success_count} members.\nFailed to import: {fail_count} members."
                if failed_ids:
                    summary_message += f"\n\nFailed PKF IDs:\n" + ",\n".join(failed_ids)
                
                messagebox.showinfo("Import Summary", summary_message)
                
                if self.on_save_callback:
                    self.on_save_callback()
            elif result_type == "import_error":
                error_message = data
                messagebox.showerror("Import Error", f"An error occurred during the import process: {error_message}")
        except Empty:
            pass
        finally:
            self.after(100, self._process_import_queue)

    def _process_save_queue(self):
        """Processes results from the background save worker thread."""
        try:
            result_type, data = self.save_queue.get_nowait()
            self.save_button.configure(state="normal", text="Update Member" if self.member_data else "Save Member")

            if result_type == "save_success":
                member_name = data
                messagebox.showinfo("Success", f"Member '{member_name}' has been saved/updated successfully.")
                
                if self.on_save_callback:
                    self.on_save_callback()
                
                if isinstance(self.master, ctk.CTkToplevel):
                    self.master.destroy()
                elif not self.member_data:
                    self.clear_form()

            elif result_type == "save_error":
                error_message = data
                messagebox.showerror("Database Error", f"Failed to save member: {error_message}")
        except Empty:
            pass
        finally:
            self.after(100, self._process_save_queue)

    def _add_achievement_to_tree(self):
        ach_type = self.ach_type_entry.get()
        ach_place = self.ach_place_entry.get()
        ach_date = self.ach_date_entry.get()
        if ach_type and ach_place and ach_date:
            self.ach_tree.insert("", "end", values=(ach_type, ach_place, ach_date))
            self.ach_type_entry.delete(0, "end")
            self.ach_place_entry.delete(0, "end")
            self.ach_date_entry.delete(0, "end")
        else:
            messagebox.showwarning("Warning", "Please fill all three achievement fields (Type, Place, Date).")

    def _update_selected_categories_display(self):
        """Updates the textbox showing the currently selected categories."""
        if 'Player' not in self.specific_widgets:
            return

        player_widgets = self.specific_widgets['Player']['widgets']
        if 'category_checkboxes' not in player_widgets:
            return

        selected_categories = [name for name, chk in player_widgets['category_checkboxes'].items() if chk.get() == 1]

        # Update the display textbox
        self.selected_categories_display.configure(state="normal")
        self.selected_categories_display.delete("1.0", "end")
        self.selected_categories_display.insert("1.0", "\n".join(selected_categories))
        self.selected_categories_display.configure(state="disabled")

    def _import_from_excel(self):
        filepath = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not filepath:
            return # User cancelled

        messagebox.showinfo("Importing", "Importing members from Excel in the background. You will be notified upon completion.")
        
        # Run the heavy lifting in a separate thread
        thread = threading.Thread(target=self._import_from_excel_worker, args=(filepath,), daemon=True)
        thread.start()

    def _save_member_worker(self, data, is_update, member_id=None):
        """Worker function to save or update member data in the background."""
        try:
            if is_update:
                update_member(member_id, data)
            else:
                add_member(data)
            self.save_queue.put(("save_success", data['full_name']))
        except Exception as e:
            self.save_queue.put(("save_error", str(e)))

    def _import_from_excel_worker(self, filepath):
        """Worker function to handle reading the Excel file and inserting into DB."""
        try:
            workbook = load_workbook(filepath)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]
            
            main_keys = {
                'pkf_id', 'id_number', 'full_name', 'full_name_ar', 'dob', 'phone', 'email',
                'role', 'expiry_date', 'current_belt', 'notes', 'admin_title',
                'passport_number', 'passport_expiry_date'
            }
            success_count = 0
            fail_count = 0
            failed_ids = []

            for row_index in range(2, sheet.max_row + 1):
                try:
                    member_data = {}
                    specific_data = {}
                    row_values = [cell.value for cell in sheet[row_index]]
                    row_dict = dict(zip(headers, row_values))

                    if not row_dict.get('pkf_id'):
                        continue

                    for key, value in row_dict.items():
                        if value is None: continue
                        if key in main_keys:
                            member_data[key] = value
                        else:
                            if isinstance(value, (int, float)) and value == 1:
                                specific_data[key] = True
                            else:
                                specific_data[key] = value
                    
                    club_name_from_excel = row_dict.get('club_name')
                    if club_name_from_excel:
                        club_id = self.clubs_map.get(club_name_from_excel)
                        if club_id:
                            member_data['club_id'] = club_id

                    member_data['specific_data'] = specific_data

                    add_member(member_data)
                    success_count += 1
                except Exception as e:
                    fail_count += 1
                    failed_ids.append(str(member_data.get('pkf_id', 'Unknown')))
                    print(f"Failed to import member {member_data.get('pkf_id')}: {e}")
            
            # Put the final results into the queue for the main thread to display
            self.import_queue.put(("import_finished", (success_count, fail_count, failed_ids)))
        except Exception as e:
            self.import_queue.put(("import_error", str(e)))

    def _download_excel_template(self):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")],
            initialfile="pkf_import_template.xlsx"
        )
        if not filepath:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Members Import"

        # Define all possible headers
        headers = [
            'pkf_id', 'id_number', 'full_name', 'full_name_ar', 'dob', 'phone', 'email',
            'passport_number', 'passport_expiry_date',
            'club_name',  # For matching during import
            'role', 'expiry_date', 'current_belt', 'notes', 'admin_title',
            'kata_check', 'kata_individual', 'kata_team', 'kumite_check', 'kumite_individual', 'kumite_team', 'weight', 'nat_rank', 'nat_rank_points', 'int_rank', 'int_rank_points', 'additional_details',
            'license_date', 'ref_kata_check', 'ref_kata_judge_b', 'ref_kata_judge_a', 'ref_kumite_check', 'ref_kumite_judge_b', 'ref_kumite_judge_a', 'ref_kumite_referee_b', 'ref_kumite_referee_a', 'ref_kumite_general', 'refereeing_achievements'
        ]
        ws.append(headers)
        wb.save(filepath)
        messagebox.showinfo("Success", f"Excel template saved to:\n{filepath}")

    def _create_multi_attachment_field(self, master, text, category_key):
        """Creates a standardized frame for uploading multiple files for a given category."""
        main_frame = ctk.CTkFrame(master)
        main_frame.grid_columnconfigure(0, weight=1)

        # --- Header with Upload Button ---
        header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        header_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=(5, 0))
        header_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(header_frame, text=text).grid(row=0, column=0, sticky="w")
        upload_button = ctk.CTkButton(header_frame, text="Upload Files", width=120, command=lambda: self._upload_multiple_files(category_key))
        upload_button.grid(row=0, column=1, sticky="e")

        # --- Treeview for file list ---
        tree_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        tree = Treeview(tree_frame, columns=("filename",), show="headings", height=3)
        tree.heading("filename", text="File Name")
        tree.column("filename", anchor="w")
        tree.grid(row=0, column=0, sticky="nsew")
        bind_mouse_wheel(tree)

        # --- Scrollbar for Treeview ---
        scrollbar = ctk.CTkScrollbar(tree_frame, command=tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=scrollbar.set)

        # Store the tree widget for later access
        self.multi_attachments[category_key] = {'tree': tree}

        # --- Remove Button ---
        remove_button = ctk.CTkButton(main_frame, text="Remove Selected File", command=lambda: self._remove_attachment_from_tree(category_key))
        remove_button.grid(row=2, column=0, sticky="w", padx=5, pady=(0, 5))

        return main_frame

    def _create_entry_fields(self, master, fields, start_row=0):
        master.grid_columnconfigure(1, weight=1)
        for i, (key, placeholder) in enumerate(fields):
            row = start_row + i
            label_text = placeholder.replace('*', '').strip()
            ctk.CTkLabel(master, text=f"{label_text}:").grid(row=row, column=0, padx=10, pady=5, sticky="w")
            
            if "Date" in placeholder or "Expiry" in placeholder:
                entry = DateEntry(master, placeholder_text=placeholder)
            else:
                entry = ctk.CTkEntry(master, placeholder_text=placeholder)
            entry.grid(row=row, column=1, padx=10, pady=5, sticky="ew")
            self.entries[key] = entry

    def _create_player_fields(self, master):
        frame = ctk.CTkFrame(master, fg_color="transparent")
        frame.grid_columnconfigure(1, weight=1) # Make entry columns expandable
        frame.grid_columnconfigure(3, weight=1)
        self.specific_widgets['Player'] = {'frame': frame, 'widgets': {}}
        widgets = self.specific_widgets['Player']['widgets']
        
        # --- Participation Checkboxes ---
        ctk.CTkLabel(frame, text="Kata:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        widgets['kata_check'] = ctk.CTkCheckBox(frame, text="")
        widgets['kata_check'].grid(row=0, column=1, padx=0, pady=5, sticky="w")
        widgets['kata_individual'] = ctk.CTkCheckBox(frame, text="Individual")
        widgets['kata_individual'].grid(row=0, column=2, padx=5, pady=5, sticky="w")
        widgets['kata_team'] = ctk.CTkCheckBox(frame, text="Team")
        widgets['kata_team'].grid(row=0, column=3, padx=5, pady=5, sticky="w")

        ctk.CTkLabel(frame, text="Kumite:", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, padx=10, pady=5, sticky="w")
        widgets['kumite_check'] = ctk.CTkCheckBox(frame, text="")
        widgets['kumite_check'].grid(row=1, column=1, padx=0, pady=5, sticky="w")
        widgets['kumite_individual'] = ctk.CTkCheckBox(frame, text="Individual")
        widgets['kumite_individual'].grid(row=1, column=2, padx=5, pady=5, sticky="w")
        widgets['kumite_team'] = ctk.CTkCheckBox(frame, text="Team")
        widgets['kumite_team'].grid(row=1, column=3, padx=5, pady=5, sticky="w")

        # --- Weight and Ranking ---
        weight_options = [str(i) for i in range(10, 151)] # Weights from 10kg to 150kg
        widgets['weight'] = (ctk.CTkLabel(frame, text="Weight (kg):"), ctk.CTkComboBox(frame, values=weight_options, command=self._update_age_and_categories))
        widgets['weight'][0].grid(row=2, column=0, padx=10, pady=5, sticky="w")
        widgets['weight'][1].grid(row=2, column=1, columnspan=3, padx=10, pady=5, sticky="ew")
        widgets['weight'][1].set("") # Set initial value to empty
        widgets['weight'][1].bind("<FocusOut>", self._update_age_and_categories)
        widgets['weight'][1].bind("<Return>", self._update_age_and_categories)

        widgets['nat_rank'] = (ctk.CTkLabel(frame, text="National Rank:"), ctk.CTkEntry(frame))
        widgets['nat_rank_points'] = (ctk.CTkLabel(frame, text="Points:"), ctk.CTkEntry(frame))
        widgets['nat_rank'][0].grid(row=3, column=0, padx=10, pady=5, sticky="w")
        widgets['nat_rank'][1].grid(row=3, column=1, padx=10, pady=5, sticky="ew")
        widgets['nat_rank_points'][0].grid(row=3, column=2, padx=10, pady=5, sticky="w")
        widgets['nat_rank_points'][1].grid(row=3, column=3, padx=10, pady=5, sticky="ew")

        widgets['int_rank'] = (ctk.CTkLabel(frame, text="International Rank:"), ctk.CTkEntry(frame))
        widgets['int_rank_points'] = (ctk.CTkLabel(frame, text="Points:"), ctk.CTkEntry(frame))
        widgets['int_rank'][0].grid(row=4, column=0, padx=10, pady=5, sticky="w")
        widgets['int_rank'][1].grid(row=4, column=1, padx=10, pady=5, sticky="ew")
        widgets['int_rank_points'][0].grid(row=4, column=2, padx=10, pady=5, sticky="w")
        widgets['int_rank_points'][1].grid(row=4, column=3, padx=10, pady=5, sticky="ew")

        # --- Achievements Table ---
        ach_frame = ctk.CTkFrame(frame)
        ach_frame.grid(row=5, column=0, columnspan=4, sticky="nsew", padx=5, pady=10)
        ach_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(ach_frame, text="Achievements", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, columnspan=4, pady=(5,10))

        self.ach_type_entry = ctk.CTkEntry(ach_frame, placeholder_text="Achievement Type") # New
        self.ach_type_entry.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
        self.ach_place_entry = ctk.CTkEntry(ach_frame, placeholder_text="Place") # New
        self.ach_place_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        self.ach_date_entry = DateEntry(ach_frame, placeholder_text="Date")
        self.ach_date_entry.grid(row=1, column=2, padx=5, pady=5, sticky="ew")
        
        add_ach_button = ctk.CTkButton(ach_frame, text="Add", width=60, command=self._add_achievement_to_tree)
        add_ach_button.grid(row=1, column=3, padx=5, pady=5)

        self.ach_tree = Treeview(ach_frame, columns=("Type", "Place", "Date"), show="headings", height=4)
        self.ach_tree.heading("Type", text="Type")
        self.ach_tree.heading("Place", text="Place")
        self.ach_tree.heading("Date", text="Date")
        bind_mouse_wheel(self.ach_tree)
        self.ach_tree.grid(row=2, column=0, columnspan=4, sticky="nsew", padx=5, pady=5)

        # --- Additional Details ---
        widgets['additional_details'] = (ctk.CTkLabel(frame, text="Additional Details:"), ctk.CTkTextbox(frame, height=80))
        widgets['additional_details'][0].grid(row=6, column=0, padx=10, pady=5, sticky="nw")
        widgets['additional_details'][1].grid(row=6, column=1, columnspan=3, padx=10, pady=5, sticky="ew")

        # --- Competition Categories (New Multi-Select) ---
        category_frame = ctk.CTkFrame(frame)
        category_frame.grid(row=7, column=0, columnspan=4, sticky="nsew", padx=5, pady=10)
        category_frame.grid_columnconfigure(0, weight=1)
        category_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(category_frame, text="Eligible Categories", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.eligible_categories_frame = ctk.CTkScrollableFrame(category_frame, height=120)
        self.eligible_categories_frame.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
        bind_mouse_wheel(self.eligible_categories_frame)

        ctk.CTkLabel(category_frame, text="Selected Categories", font=ctk.CTkFont(weight="bold")).grid(row=0, column=1, padx=10, pady=5, sticky="w")
        self.selected_categories_display = ctk.CTkTextbox(category_frame, height=120, state="disabled")
        self.selected_categories_display.grid(row=1, column=1, padx=10, pady=5, sticky="nsew")

        # This will hold the checkbox widgets
        widgets['category_checkboxes'] = {}

    def _create_coach_fields(self, master):
        frame = ctk.CTkFrame(master, fg_color="transparent")
        frame.grid_columnconfigure(1, weight=1)
        self.specific_widgets['Coach'] = {'frame': frame, 'widgets': {}}
        widgets = self.specific_widgets['Coach']['widgets']

        # Helper function to create a block for each coaching level
        def create_level_block(parent, level_name_en, level_name_ar, level_key, start_row):
            header_frame = ctk.CTkFrame(parent, fg_color=("gray85", "gray25"))
            header_frame.grid(row=start_row, column=0, columnspan=2, sticky="ew", padx=5, pady=(15, 5))
            ctk.CTkLabel(header_frame, text=f"{level_name_en} Level - {level_name_ar}", font=ctk.CTkFont(weight="bold")).pack(padx=10, pady=4)

            fields = [
                (f'coach_{level_key}_degree', "Degree:", "الدرجة"),
                (f'coach_{level_key}_place', "Place:", "المكان"),
                (f'coach_{level_key}_date', "Date:", "التاريخ"),
                (f'coach_{level_key}_points', "Points:", "النقاط")
            ]
            for i, (key, label_en, label_ar) in enumerate(fields):
                row = start_row + i + 1
                label_text = f"{label_en} / {label_ar}"
                label = ctk.CTkLabel(parent, text=label_text)
                label.grid(row=row, column=0, padx=10, pady=5, sticky="w")

                if "date" in key:
                    entry = DateEntry(parent)
                else:
                    entry = ctk.CTkEntry(parent)
                entry.grid(row=row, column=1, padx=10, pady=5, sticky="ew")
                widgets[key] = (label, entry)
            return start_row + len(fields) + 1

        # Create blocks for each level
        next_row = create_level_block(frame, "National", "وطني", "national", 0)
        next_row = create_level_block(frame, "Asian", "آسيوي", "asian", next_row)
        next_row = create_level_block(frame, "International", "دولي", "international", next_row)

        widgets['additional_details'] = (ctk.CTkLabel(frame, text="Additional Details:"), ctk.CTkTextbox(frame, height=80))
        widgets['additional_details'][0].grid(row=next_row, column=0, padx=10, pady=5, sticky="nw")
        widgets['additional_details'][1].grid(row=next_row, column=1, padx=10, pady=5, sticky="ew")

    def _create_referee_fields(self, master):
        frame = ctk.CTkFrame(master, fg_color="transparent")
        frame.grid_columnconfigure(1, weight=1)
        self.specific_widgets['Referee'] = {'frame': frame, 'widgets': {}}
        widgets = self.specific_widgets['Referee']['widgets']

        # Helper function to create a block for each level (National, Asian, Intl)
        def create_level_block(parent, discipline_key, level_name_en, level_name_ar, level_key, start_row):
            header_frame = ctk.CTkFrame(parent, fg_color=("gray85", "gray25"))
            header_frame.grid(row=start_row, column=0, columnspan=2, sticky="ew", padx=5, pady=(10, 5))
            ctk.CTkLabel(header_frame, text=f"{level_name_en} Level - {level_name_ar}", font=ctk.CTkFont(weight="bold")).pack(padx=10, pady=4)

            fields = [
                (f'ref_{discipline_key}_{level_key}_degree', "Degree/Rank:", "الدرجة/التصنيف"),
                (f'ref_{discipline_key}_{level_key}_place', "Place:", "المكان"),
                (f'ref_{discipline_key}_{level_key}_date', "Date:", "التاريخ"),
                (f'ref_{discipline_key}_{level_key}_points', "Points:", "النقاط")
            ]
            for i, (key, label_en, label_ar) in enumerate(fields):
                row = start_row + i + 1
                label_text = f"{label_en} / {label_ar}"
                label = ctk.CTkLabel(parent, text=label_text)
                label.grid(row=row, column=0, padx=10, pady=5, sticky="w")

                if "date" in key:
                    entry = DateEntry(parent)
                else:
                    entry = ctk.CTkEntry(parent)
                entry.grid(row=row, column=1, padx=10, pady=5, sticky="ew")
                widgets[key] = (label, entry)
            return start_row + len(fields) + 1

        # Helper function to create a discipline section (Kata or Kumite)
        def create_discipline_section(parent, discipline_name_en, discipline_name_ar, discipline_key, start_row):
            discipline_header = ctk.CTkLabel(parent, text=f"{discipline_name_en} Classification - تصنيف {discipline_name_ar}", font=ctk.CTkFont(size=16, weight="bold"))
            discipline_header.grid(row=start_row, column=0, columnspan=2, pady=(20, 5), sticky="w")
            
            next_row = create_level_block(parent, discipline_key, "National", "وطني", "national", start_row + 1)
            next_row = create_level_block(parent, discipline_key, "Asian", "آسيوي", "asian", next_row)
            next_row = create_level_block(parent, discipline_key, "International", "دولي", "international", next_row)
            return next_row

        # Create sections for Kata and Kumite
        next_row = create_discipline_section(frame, "Kata", "الكاتا", "kata", 0)
        next_row = create_discipline_section(frame, "Kumite", "الكوميتيه", "kumite", next_row)

        # --- Refereeing Achievements ---
        widgets['refereeing_achievements'] = (ctk.CTkLabel(frame, text="Refereeing Achievements:"), ctk.CTkTextbox(frame, height=80))
        widgets['refereeing_achievements'][0].grid(row=next_row, column=0, padx=10, pady=5, sticky="nw")
        widgets['refereeing_achievements'][1].grid(row=next_row, column=1, padx=10, pady=5, sticky="ew")

    def _create_admin_fields(self, master):
        frame = ctk.CTkFrame(master, fg_color="transparent")
        frame.grid_columnconfigure(1, weight=1)
        self.specific_widgets['Admin'] = {'frame': frame, 'widgets': {}}
        widgets = self.specific_widgets['Admin']['widgets']

        # --- Admin Title ---
        widgets['admin_title'] = (ctk.CTkLabel(frame, text="Admin Title:"), ctk.CTkEntry(frame, placeholder_text="e.g., Federation President, Board Member"))
        widgets['admin_title'][0].grid(row=0, column=0, padx=10, pady=5, sticky="w")
        widgets['admin_title'][1].grid(row=0, column=1, padx=10, pady=5, sticky="ew")

    def _update_role_fields(self, selected_role):
        # Hide all specific widgets first
        for role_info in self.specific_widgets.values():
            role_info['frame'].pack_forget()

        if selected_role in self.specific_widgets:
            self.specific_widgets[selected_role]['frame'].pack(fill="x", expand=True, padx=5, pady=5)

        # Show/hide role-specific attachment sections
        is_coach = selected_role == "Coach"
        is_referee = selected_role == "Referee"

        if 'coach_certs' in self.multi_attachments:
            if is_coach:
                self.multi_attachments['coach_certs']['frame'].grid()
            else:
                self.multi_attachments['coach_certs']['frame'].grid_remove()

        if 'referee_certs' in self.multi_attachments:
            if is_referee:
                self.multi_attachments['referee_certs']['frame'].grid()
            else:
                self.multi_attachments['referee_certs']['frame'].grid_remove()

    def _upload_personal_photo(self):
        filepath = filedialog.askopenfilename(title="Select Personal Photo", filetypes=(("Image files", "*.jpg *.jpeg *.png"), ("All files", "*.*")))
        if not filepath:
            return
        
        pkf_id = self.entries['pkf_id'].get()
        if not pkf_id:
            messagebox.showwarning("Warning", "Please enter a Membership No. (PKF ID) before uploading files.")
            return

        # Sanitize pkf_id for use in directory name
        safe_pkf_id = "".join(c for c in pkf_id if c.isalnum() or c in ('-', '_')).rstrip()
        member_assets_dir = os.path.join("assets", "member_files", safe_pkf_id)
        os.makedirs(member_assets_dir, exist_ok=True)
        
        # Use a descriptive filename
        original_filename = os.path.basename(filepath)
        destination_filename = f"personal_photo_{original_filename}"
        destination_path = os.path.join(member_assets_dir, destination_filename)
        
        shutil.copy(filepath, destination_path)
        self.photo_path_var.set(destination_path)
        self.photo_path_label.configure(text=os.path.basename(destination_path), text_color=("black", "white"))

    def _upload_multiple_files(self, category_key):
        filepaths = filedialog.askopenfilenames(title=f"Select File(s) for {category_key.replace('_', ' ').title()}")
        if not filepaths:
            return

        pkf_id = self.entries['pkf_id'].get()
        if not pkf_id:
            messagebox.showwarning("Warning", "Please enter a Membership No. (PKF ID) before uploading files.")
            return

        safe_pkf_id = "".join(c for c in pkf_id if c.isalnum() or c in ('-', '_')).rstrip()
        member_assets_dir = os.path.join("assets", "member_files", safe_pkf_id)
        os.makedirs(member_assets_dir, exist_ok=True)

        tree = self.multi_attachments[category_key]['tree']
        for filepath in filepaths:
            original_filename = os.path.basename(filepath)
            destination_filename = f"{category_key}_{original_filename}"
            destination_path = os.path.join(member_assets_dir, destination_filename)
            
            counter = 1
            while os.path.exists(destination_path):
                name, ext = os.path.splitext(destination_filename)
                destination_filename = f"{name}_{counter}{ext}"
                destination_path = os.path.join(member_assets_dir, destination_filename)
                counter += 1

            shutil.copy(filepath, destination_path)
            tree.insert("", "end", iid=destination_path, values=(os.path.basename(destination_path),))

    def _remove_attachment_from_tree(self, category_key):
        tree = self.multi_attachments[category_key]['tree']
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select a file to remove.")
            return

        # Ask for confirmation for all selected files at once
        if messagebox.askyesno("Confirm Deletion", f"Are you sure you want to remove {len(selected_items)} selected file(s) from the list and delete them from disk? This action cannot be undone."):
            for item_id in selected_items:
                # The item_id (iid) is the full path to the file, which we set during upload
                filepath = item_id
                try:
                    if os.path.exists(filepath):
                        os.remove(filepath)
                    
                    # Remove from treeview
                    tree.delete(item_id)
                except Exception as e:
                    messagebox.showerror("Error", f"Could not delete file {os.path.basename(filepath)}: {e}")

    def _save_member(self):
        # Collect data from all entries
        data = {}
        for key, widget in self.entries.items():
            if isinstance(widget, (ctk.CTkEntry, ctk.CTkOptionMenu, ctk.CTkComboBox)):
                data[key] = widget.get()
            elif isinstance(widget, ctk.CTkTextbox):
                data[key] = widget.get("1.0", "end-1c")
            elif isinstance(widget, DateEntry):
                data[key] = widget.get()

        data['photo_path'] = self.photo_path_var.get()

        # Collect belt history data from Treeview
        belts_data = []
        for item_id in self.belt_tree.get_children():
            values = self.belt_tree.item(item_id, 'values')
            belts_data.append({'belt_name': values[0], 'date_obtained': values[1], 'source': values[2]})
        data['belts'] = belts_data

        # Collect achievement data from Treeview if player
        if data.get("role") == "Player":
            achievements_data = []
            for item_id in self.ach_tree.get_children():
                values = self.ach_tree.item(item_id, 'values')
                achievements_data.append({'achievement_type': values[0], 'place': values[1], 'date': values[2]})
            data['achievements'] = achievements_data

        # Basic validation
        if not all([data.get("full_name"), data.get("pkf_id"), data.get("role")]):
            messagebox.showerror("Error", "Full Name, Membership No. (PKF ID), and Role are required.")
            return

        # --- Club Selection Validation ---
        selected_club_name = self.club_combobox.get()
        if not selected_club_name or selected_club_name in ["Select a club", "No clubs available"]:
            messagebox.showerror("Validation Error", "'Club' is a required field.")
            return
        # Add club_id to the data to be saved
        data['club_id'] = self.clubs_map.get(selected_club_name)

        # Collect role-specific data
        data["specific_data"] = {}
        role = data.get("role")
        if role in self.specific_widgets:
            role_widgets = self.specific_widgets[role]['widgets']
            for key, widget_item in role_widgets.items():
                if isinstance(widget_item, tuple): # Label, Entry
                    entry = widget_item[1]
                    if isinstance(entry, (ctk.CTkEntry, ctk.CTkComboBox)):
                        data["specific_data"][key] = entry.get()
                    elif isinstance(entry, ctk.CTkTextbox):
                        data["specific_data"][key] = entry.get("1.0", "end-1c")
                elif isinstance(widget_item, ctk.CTkCheckBox): # Checkbox
                    data["specific_data"][key] = widget_item.get() == 1 # Store as boolean

        # Handle multi-select categories for Player role
        if role == "Player":
            selected_categories = [cat for cat, chk in self.specific_widgets['Player']['widgets']['category_checkboxes'].items() if chk.get() == 1]
            data["specific_data"]['competition_categories'] = selected_categories

        # Collect multi-attachment data
        for key, attachment_info in self.multi_attachments.items():
            tree = attachment_info['tree']
            file_paths = tree.get_children() # Item IDs are the full paths
            if file_paths:
                data["specific_data"][key] = file_paths

        # Disable button and show message
        self.save_button.configure(state="disabled", text="Saving...")
        self.update_idletasks()

        is_update = self.member_data is not None
        member_id = self.member_data['id'] if is_update else None

        # Run save in a background thread
        thread = threading.Thread(target=self._save_member_worker, args=(data, is_update, member_id), daemon=True)
        thread.start()

    def clear_form(self, is_prefill=False):
        """Clears all input fields in the form."""
        for key, widget in self.entries.items():
            # If prefilling, we don't clear pkf_id here, it will be set by _prefill_for_new_role
            if key == 'pkf_id':
                continue
            if isinstance(widget, ctk.CTkEntry):
                widget.delete(0, 'end')
            elif isinstance(widget, ctk.CTkTextbox):
                widget.delete("1.0", "end")
            elif isinstance(widget, ctk.CTkComboBox):
                widget.set("")
            elif isinstance(widget, DateEntry):
                widget.delete(0, 'end')
        
        # Reset photo attachment
        self.photo_path_var.set("")
        self.photo_path_label.configure(text="No file selected.", text_color="gray")

        # Clear multi-attachment trees
        for attachment_info in self.multi_attachments.values():
            for item in attachment_info['tree'].get_children():
                attachment_info['tree'].delete(item)

        for role_info in self.specific_widgets.values():
            for widget_item in role_info['widgets'].values():
                if isinstance(widget_item, tuple):
                    entry = widget_item[1]
                    if isinstance(entry, ctk.CTkEntry): entry.delete(0, 'end')
                    elif isinstance(entry, ctk.CTkTextbox): entry.delete('1.0', 'end')
                    elif isinstance(entry, DateEntry): entry.delete(0, 'end')

                elif isinstance(widget_item, ctk.CTkComboBox):  # Clear comboboxes
                    widget_item.set("")
                elif isinstance(widget_item, ctk.CTkCheckBox):
                    widget_item.deselect()

        self.role_menu.set("Player")
        # Clear belt history table
        for item in self.belt_tree.get_children():
            self.belt_tree.delete(item)
        # Clear achievement history table
        if hasattr(self, 'ach_tree'):
            for item in self.ach_tree.get_children():
                self.ach_tree.delete(item)
        # Clear referee classifications table
        if hasattr(self, 'ref_class_tree'):
            for item in self.ref_class_tree.get_children():
                self.ref_class_tree.delete(item) # This line is now redundant but safe to keep for old versions
        self._update_age_and_categories() # Clear/reset categories
        self.update_club_list() # Always update club list

    def _populate_club_selection(self, data_source=None):
        """Populates the club combobox based on the provided data source (or self.member_data)."""
        if data_source is None:
            data_source = self.member_data
        
        club_id = data_source.get('club_id')
        if club_id:
            club_name = next((name for name, c_id in self.clubs_map.items() if c_id == club_id), None)
            if club_name:
                self.club_combobox.set(club_name)


    def populate_form(self):
        """Fills the form with existing member data for editing."""
        if not self.member_data:
            return

        # Populate main entries
        for key, widget in self.entries.items():
            value = self.member_data.get(key, '')
            if value is None: value = '' # Ensure value is not None

            if isinstance(widget, ctk.CTkEntry):
                widget.insert(0, str(value))
            elif isinstance(widget, (ctk.CTkOptionMenu, ctk.CTkComboBox)):
                widget.set(str(value))
            elif isinstance(widget, DateEntry):
                widget.set(str(value))
            elif isinstance(widget, ctk.CTkTextbox):
                widget.insert("1.0", str(value))

        # Disable PKF ID editing
        self.entries['pkf_id'].configure(state="disabled")

        self._populate_club_selection()

        # Populate photo path
        photo_path = self.member_data.get('photo_path')
        if photo_path and os.path.exists(photo_path):
            self.photo_path_var.set(photo_path)
            self.photo_path_label.configure(text=os.path.basename(photo_path), text_color=("black", "white"))

        # Populate belt history
        belt_history = get_belts(self.member_data.get('pkf_id'))
        for item in self.belt_tree.get_children():
            self.belt_tree.delete(item)
        for belt in belt_history:
            self.belt_tree.insert("", "end", values=(belt['belt_name'], belt['date_obtained'], belt['source']))

        # Populate achievement history if player
        if self.member_data.get('role') == 'Player':
            ach_history = get_achievements(self.member_data.get('pkf_id'))
            for item in self.ach_tree.get_children():
                self.ach_tree.delete(item)
            for ach in ach_history:
                self.ach_tree.insert("", "end", values=(ach['achievement_type'], ach['place'], ach['date']))

        # Populate specific data
        try:
            specific_data = json.loads(self.member_data.get('specific_data', '{}'))
            role = self.member_data.get('role')
            if role and role in self.specific_widgets:
                role_widgets = self.specific_widgets[role]['widgets']
                for key, widget_item in role_widgets.items():
                    value = specific_data.get(key)
                    if value is None: continue

                    if isinstance(widget_item, tuple): # (Label, Entry)
                        entry = widget_item[1]
                        if isinstance(entry, ctk.CTkEntry):
                            entry.insert(0, str(value))
                        elif isinstance(entry, ctk.CTkTextbox):
                            entry.insert("1.0", str(value))
                        elif isinstance(entry, ctk.CTkComboBox):
                            entry.set(str(value))
                        elif isinstance(entry, DateEntry):
                            entry.set(str(value))
                    elif isinstance(widget_item, ctk.CTkCheckBox): # Checkbox
                        if value:
                            widget_item.select()
                        else:
                            widget_item.deselect()
            self._update_age_and_categories() # Update categories and age after populating

            # Now, handle populating the multi-select categories after they have been created
            if role == 'Player':
                player_widgets = self.specific_widgets[role]['widgets']
                saved_categories = specific_data.get('competition_categories', [])
                if saved_categories and 'category_checkboxes' in player_widgets:
                    for category_name, checkbox in player_widgets['category_checkboxes'].items():
                        if category_name in saved_categories:
                            checkbox.select()
                    self._update_selected_categories_display() # Update the display textbox
            
            # Populate multi-attachments
            for key, attachment_info in self.multi_attachments.items():
                if key in specific_data:
                    tree = attachment_info['tree']
                    file_paths = specific_data[key]
                    if isinstance(file_paths, list):
                        for path in file_paths:
                            if os.path.exists(path):
                                tree.insert("", "end", iid=path, values=(os.path.basename(path),))
        except (json.JSONDecodeError, TypeError):
            print("Could not parse or populate specific_data.")
        
        # Update role fields display at the end, after all data is populated
        self._update_role_fields(self.member_data.get('role', 'Player'))